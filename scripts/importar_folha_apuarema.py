"""
importar_folha_apuarema.py
--------------------------
Lê a aba "Folha de pagamento normal" do Excel
  C:\\Users\\pichau\\OneDrive\\Documentos\\Folha de Pagamento.xlsx
e cria um novo registro de folha (status=rascunho) no Supabase,
populando folha_funcionarios.

Uso:
    python scripts/importar_folha_apuarema.py
    python scripts/importar_folha_apuarema.py --arquivo "outro/caminho.xlsx"
"""

import argparse
import os
import shutil
import sys
import tempfile
from datetime import datetime

import openpyxl
from dotenv import load_dotenv
from supabase import create_client

# ── Configurações ──────────────────────────────────────────────────────────────

ARQUIVO_PADRAO = r"C:\Users\pichau\OneDrive\Documentos\Folha de Pagamento.xlsx"
ABA = "Folha de pagamento normal"
OBRA = "Creche apuarema"

# Índices de coluna (0-based)
COL_NOME      = 0   # A
COL_PIX       = 1   # B
COL_NOME_CONTA = 2  # C
COL_SERVICO   = 3   # D
COL_ETAPA     = 4   # E
COL_DIARIAS   = 5   # F
COL_VALOR     = 6   # G

LINHA_INICIO_DADOS = 4   # linha do Excel onde começam os funcionários
ZEROS_CONSECUTIVOS_PARA_PARAR = 5

# Mapeamento de nomes de serviço da planilha → nome exato cadastrado em folha_regras
SERVICO_MAP = {
    "CLT": "CLT'S",
}


# ── Helpers ────────────────────────────────────────────────────────────────────

def _str(v) -> str | None:
    """Converte valor para string limpa, ou None se vazio."""
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def _float(v) -> float | None:
    """Converte para float, ou None se não for numérico."""
    if v is None:
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def ler_planilha(arquivo: str) -> tuple[str, list[dict]]:
    """
    Lê o arquivo Excel e retorna (quinzena_str, lista_de_funcionarios).
    Copia para tmp para evitar PermissionError com arquivo aberto no Excel.
    """
    tmp = tempfile.mktemp(suffix=".xlsx")
    shutil.copy2(arquivo, tmp)
    try:
        wb = openpyxl.load_workbook(tmp, data_only=True)
    finally:
        os.unlink(tmp)

    if ABA not in wb.sheetnames:
        print(f"[ERRO] Aba '{ABA}' não encontrada. Abas disponíveis: {wb.sheetnames}")
        sys.exit(1)

    ws = wb[ABA]

    quinzena = datetime.today().strftime("%Y-%m-%d")

    funcionarios = []
    zeros_consecutivos = 0

    for row_idx, row in enumerate(ws.iter_rows(min_row=LINHA_INICIO_DADOS, values_only=True), start=LINHA_INICIO_DADOS):
        valor = _float(row[COL_VALOR] if len(row) > COL_VALOR else None)

        # Conta zeros consecutivos para detectar fim da planilha
        if valor is None or valor == 0:
            zeros_consecutivos += 1
            if zeros_consecutivos >= ZEROS_CONSECUTIVOS_PARA_PARAR:
                break
            continue
        zeros_consecutivos = 0

        nome      = _str(row[COL_NOME] if len(row) > COL_NOME else None)
        pix_raw   = row[COL_PIX] if len(row) > COL_PIX else None
        nome_conta = _str(row[COL_NOME_CONTA] if len(row) > COL_NOME_CONTA else None)
        servico_raw = _str(row[COL_SERVICO] if len(row) > COL_SERVICO else None)
        servico   = SERVICO_MAP.get(servico_raw, servico_raw) if servico_raw else None
        etapa     = _str(row[COL_ETAPA] if len(row) > COL_ETAPA else None)
        diarias   = _float(row[COL_DIARIAS] if len(row) > COL_DIARIAS else None)

        # Normaliza PIX (int vira string, remove espaços)
        pix = _str(str(int(pix_raw)) if isinstance(pix_raw, float) and pix_raw == int(pix_raw)
                   else str(pix_raw) if pix_raw is not None else None)

        # Ignora linha sem identificação alguma
        if not nome and not pix and not nome_conta:
            continue

        # Funcionários sem diárias (CLT, cargos fixos) têm valor individual.
        # Usamos valor_fixo para que o frontend preserve o valor mesmo sem regra cadastrada.
        valor_fixo = round(valor, 2) if not diarias else None

        funcionarios.append({
            "nome":       nome,
            "pix":        pix,
            "nome_conta": nome_conta,
            "servico":    servico,
            "etapa":      etapa,
            "diarias":    diarias,
            "valor":      round(valor, 2),
            "valor_fixo": valor_fixo,
        })

    return quinzena, funcionarios


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Importa folha de pagamento para o Supabase.")
    parser.add_argument("--arquivo", default=ARQUIVO_PADRAO, help="Caminho para o arquivo Excel")
    args = parser.parse_args()

    # Carrega .env do projeto
    load_dotenv()
    url = os.getenv("SUPABASE_URL")
    key = os.getenv("SUPABASE_SERVICE_KEY") or os.getenv("SUPABASE_KEY")
    if not url or not key:
        print("[ERRO] SUPABASE_URL e SUPABASE_SERVICE_KEY precisam estar no .env")
        sys.exit(1)

    sb = create_client(url, key)

    print(f"Lendo: {args.arquivo}")
    quinzena, funcionarios = ler_planilha(args.arquivo)

    if not funcionarios:
        print("[ERRO] Nenhum funcionário válido encontrado na planilha.")
        sys.exit(1)

    print(f"Data da quinzena: {quinzena}")
    print(f"Funcionários encontrados: {len(funcionarios)}")

    # ── Verifica se já existe folha para (obra, quinzena) ─────────────────────
    existente = sb.table("folhas").select("id, status").eq("obra", OBRA).eq("quinzena", quinzena).execute()
    folha_criada_agora = False

    if existente.data:
        folha_existente = existente.data[0]
        folha_id = folha_existente["id"]

        if folha_existente["status"] == "fechada":
            print(f"[ERRO] Já existe uma folha FECHADA para {OBRA} em {quinzena} (ID {folha_id}). Nada foi alterado.")
            sys.exit(1)

        # Verifica se a folha rascunho já tem funcionários
        tem_func = sb.table("folha_funcionarios").select("id").eq("folha_id", folha_id).limit(1).execute()
        if tem_func.data:
            print(f"[ERRO] Já existe uma folha (ID {folha_id}) com funcionários para {OBRA} em {quinzena}.")
            print("       Delete os funcionários ou a folha no banco antes de reimportar.")
            sys.exit(1)

        print(f"Reusando folha existente (ID {folha_id}, status={folha_existente['status']})")
    else:
        # Cria nova folha
        folha_res = sb.table("folhas").insert({
            "obra":     OBRA,
            "quinzena": quinzena,
            "status":   "rascunho",
        }).execute()

        if not folha_res.data:
            print("[ERRO] Falha ao criar registro em 'folhas'.")
            sys.exit(1)

        folha_id = folha_res.data[0]["id"]
        folha_criada_agora = True
        print(f"Folha criada com ID: {folha_id}")

    # ── Insere funcionários individualmente (erros não travam os demais) ──────
    importados = 0
    erros = []

    for f in funcionarios:
        try:
            sb.table("folha_funcionarios").insert({**f, "folha_id": folha_id}).execute()
            importados += 1
        except Exception as e:
            erros.append({"nome": f.get("nome", "?"), "erro": str(e)})

    # Se nenhum funcionário foi inserido e a folha foi criada agora, faz rollback
    if importados == 0:
        if folha_criada_agora:
            sb.table("folhas").delete().eq("id", folha_id).execute()
            print(f"[ERRO] Nenhum funcionário inserido. Folha ID {folha_id} removida (rollback).")
        else:
            print("[ERRO] Nenhum funcionário inserido.")
        sys.exit(1)

    print(f"\n✓ Importação concluída!")
    print(f"  Folha ID  : {folha_id}")
    print(f"  Obra      : {OBRA}")
    print(f"  Quinzena  : {quinzena}")
    print(f"  Importados: {importados} funcionários")
    if erros:
        print(f"  Erros ({len(erros)}):")
        for e in erros:
            print(f"    - {e['nome']}: {e['erro']}")
    print(f"\nAcesse o frontend em http://localhost:8080/folha/ para revisar e fechar a folha.")


if __name__ == "__main__":
    main()
