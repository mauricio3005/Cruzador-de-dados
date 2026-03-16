"""
sync_to_excel.py
─────────────────────────────────────────────────────────────────────────────
Traz para o Excel as despesas cadastradas diretamente no Supabase (via
dashboard) que ainda não existem na aba "C Despesas".

Detecção de duplicatas: OBRA + ETAPA + VALOR TOTAL + DATA
(evita falsos positivos quando FORNECEDOR ou DESCRIÇÃO estão em branco)

Nunca sobrescreve nem apaga linhas existentes — apenas acrescenta.

Fluxo recomendado sempre que for rodar o app.py:
    1. python sync_to_excel.py   ← traz novidades do Supabase para o Excel
    2. python app.py             ← sincroniza Excel → Supabase
─────────────────────────────────────────────────────────────────────────────
"""

import os
import pandas as pd
from openpyxl import load_workbook
from supabase import create_client
from dotenv import load_dotenv

CAMINHO_EXCEL = r"C:\Users\mauri\OneDrive\Controle de obra\Controle de obras.xlsx"
ABA_DESPESAS  = "C Despesas"


def _normaliza(val) -> str:
    """Converte None, NaN e strings vazias para '' de forma consistente."""
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ""
    return str(val).strip().lower()


def _chave(obra, etapa, valor, data) -> str:
    """Chave de deduplicação: obra|etapa|valor|data."""
    obra_s  = _normaliza(obra)
    etapa_s = _normaliza(etapa)
    val     = round(float(valor), 2) if pd.notna(valor) and valor != "" else 0.0
    dt      = pd.to_datetime(data, errors="coerce")
    dt_s    = dt.strftime("%Y-%m-%d") if pd.notna(dt) else ""
    return f"{obra_s}|{etapa_s}|{val}|{dt_s}"


def main():
    load_dotenv()

    # ── 1. Conecta ao Supabase ────────────────────────────────────────────────
    url = os.getenv("SUPABASE_URL")
    key = os.getenv("SUPABASE_SERVICE_KEY") or os.getenv("SUPABASE_KEY") or os.getenv("SUPABASE_ANON_KEY")
    if not url or not key:
        print("❌ Credenciais do Supabase não encontradas no .env")
        return

    supabase = create_client(url, key)

    # ── 2. Lê despesas do Supabase ────────────────────────────────────────────
    print("Buscando despesas no Supabase...")
    res = supabase.table("despesas").select("*").execute()
    df_supa = pd.DataFrame(res.data)

    if df_supa.empty:
        print("Nenhuma despesa encontrada no Supabase.")
        return

    print(f"  → {len(df_supa)} registro(s) encontrado(s) no Supabase.")

    # ── 3. Lê aba "C Despesas" do Excel ──────────────────────────────────────
    print("Lendo Excel...")
    try:
        df_excel = pd.read_excel(CAMINHO_EXCEL, sheet_name=ABA_DESPESAS, header=0)
    except Exception as e:
        print(f"❌ Erro ao ler o Excel: {e}")
        print("   Verifique se o arquivo está fechado e o caminho está correto.")
        return

    df_excel.columns = df_excel.columns.str.strip()

    # Detecta colunas dinamicamente
    col_obra_excel  = next((c for c in df_excel.columns if c.strip().upper() == "OBRA"), "OBRA")
    col_etapa_excel = next((c for c in df_excel.columns if c.strip().upper() == "ETAPA"), "ETAPA")
    col_valor_excel = next(
        (c for c in df_excel.columns if "valor" in c.lower() and "total" in c.lower()), None
    )
    col_data_excel  = next((c for c in df_excel.columns if c.strip().upper() == "DATA"), "DATA")

    if col_valor_excel is None:
        print("❌ Coluna 'VALOR TOTAL' não encontrada na aba C Despesas.")
        return

    # ── 4. Monta chaves existentes no Excel ───────────────────────────────────
    chaves_excel = set()
    for _, row in df_excel.iterrows():
        chaves_excel.add(_chave(
            row.get(col_obra_excel), row.get(col_etapa_excel),
            row.get(col_valor_excel), row.get(col_data_excel)
        ))

    print(f"  → {len(chaves_excel)} chave(s) únicas já existem no Excel.")

    # ── 5. Filtra registros novos do Supabase ─────────────────────────────────
    df_supa["_chave"] = df_supa.apply(
        lambda r: _chave(r.get("OBRA"), r.get("ETAPA"), r.get("VALOR_TOTAL"), r.get("DATA")),
        axis=1
    )
    df_novos = df_supa[~df_supa["_chave"].isin(chaves_excel)].copy()

    if df_novos.empty:
        print("✅ Nenhuma despesa nova para adicionar. Excel já está atualizado.")
        return

    print(f"  → {len(df_novos)} despesa(s) nova(s) para adicionar ao Excel.")

    # ── 6. Abre workbook ──────────────────────────────────────────────────────
    try:
        wb = load_workbook(CAMINHO_EXCEL)
    except Exception as e:
        print(f"❌ Erro ao abrir o Excel para escrita: {e}")
        print("   Verifique se o arquivo está fechado.")
        return

    ws = wb[ABA_DESPESAS]

    # Descobre o número de formato da coluna DATA para preservar formatação
    cabecalho = [c.value for c in ws[1]]
    try:
        idx_data = next(
            i for i, c in enumerate(cabecalho) if c and str(c).strip().upper() == "DATA"
        )
    except StopIteration:
        idx_data = None

    data_num_format = None
    if idx_data is not None:
        # Pega o formato de uma célula existente na coluna DATA (2ª linha em diante)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=idx_data + 1, max_col=idx_data + 1):
            cell = row[0]
            if cell.value is not None and cell.number_format and cell.number_format != "General":
                data_num_format = cell.number_format
                break

    # ── 7. Acrescenta linhas ──────────────────────────────────────────────────
    adicionados = 0
    for _, row in df_novos.iterrows():
        nova_linha = []
        for col in cabecalho:
            if col is None:
                nova_linha.append(None)
                continue
            col_u = str(col).strip().upper()
            if col_u == "OBRA":
                nova_linha.append(row.get("OBRA"))
            elif col_u == "ETAPA":
                nova_linha.append(row.get("ETAPA"))
            elif col_u == "TIPO":
                nova_linha.append(row.get("TIPO"))
            elif col_u == "FORNECEDOR":
                nova_linha.append(row.get("FORNECEDOR") or None)
            elif "VALOR" in col_u and "TOTAL" in col_u:
                nova_linha.append(row.get("VALOR_TOTAL"))
            elif "VALOR" in col_u and "UNIT" in col_u:
                val_unit = row.get("VALOR_UNITARIO") or row.get("VALOR_TOTAL")
                nova_linha.append(val_unit)
            elif col_u in ("QNTD", "QTD", "QUANTIDADE"):
                nova_linha.append(row.get("QUANTIDADE") or 1)
            elif "DESCRI" in col_u:
                nova_linha.append(row.get("DESCRICAO") or None)
            elif col_u == "BANCO":
                nova_linha.append(row.get("BANCO") or None)
            elif col_u == "FORMA":
                nova_linha.append(row.get("FORMA") or None)
            elif col_u == "DESPESA":
                nova_linha.append(row.get("DESPESA") or None)
            elif col_u == "DATA":
                dt = pd.to_datetime(row.get("DATA"), errors="coerce")
                nova_linha.append(dt.to_pydatetime() if pd.notna(dt) else None)
            else:
                nova_linha.append(None)

        ws.append(nova_linha)

        # Aplica formatação de data na célula recém-adicionada
        if idx_data is not None and data_num_format:
            nova_row_idx = ws.max_row
            ws.cell(row=nova_row_idx, column=idx_data + 1).number_format = data_num_format

        adicionados += 1

    wb.save(CAMINHO_EXCEL)
    print(f"✅ {adicionados} despesa(s) adicionada(s) ao Excel com sucesso!")
    print(f"   Arquivo salvo: {CAMINHO_EXCEL}")


if __name__ == "__main__":
    main()
